# apps/payroll/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.urls import reverse
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from apps.employees.models import Employee, EmployeeAssignment
from apps.clients.models import Client
from apps.locations.models import Location
from apps.core.decorators import data_entry_or_admin_required

# ✅ Import models
from .models import PayrollRun, Payslip, PaymentVoucher, PaymentVoucherDetail, EmployeeLedger

# ✅ Import forms (Now available)
from .forms import PayrollRunForm, PaymentVoucherForm, PaymentVoucherDetailForm

# ✅ Import calculations
from .calculations import SalaryCalculator



# ============================================
# PAYROLL RUN VIEWS
# ============================================

@login_required
def run_list(request):
    """List all payroll runs"""
    payroll_runs = PayrollRun.objects.all().order_by('-month_year', '-created_at')

    search = request.GET.get('search')
    if search:
        payroll_runs = payroll_runs.filter(
            Q(run_number__icontains=search) |
            Q(client__name__icontains=search)
        )

    status = request.GET.get('status')
    if status:
        payroll_runs = payroll_runs.filter(status=status)

    paginator = Paginator(payroll_runs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'total_runs': payroll_runs.count(),
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'active': True},
        ],
    }
    return render(request, 'payroll/run_list.html', context)


@login_required
@data_entry_or_admin_required
def run_create(request):
    """Create new payroll run"""
    if request.method == 'POST':
        form = PayrollRunForm(request.POST)
        if form.is_valid():
            payroll_run = form.save(commit=False)
            payroll_run.created_by = request.user
            payroll_run.save()
            messages.success(request, f'Payroll Run {payroll_run.run_number} created successfully!')
            return redirect('payroll:run_detail', pk=payroll_run.pk)
    else:
        form = PayrollRunForm()

    context = {
        'form': form,
        'title': 'Create Payroll Run',
        'button_text': 'Create Payroll Run',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'url': '/payroll/', 'active': False},
            {'name': 'Create Payroll Run', 'active': True},
        ],
    }
    return render(request, 'payroll/run_create.html', context)


@login_required
def run_detail(request, pk):
    """Payroll run detail view"""
    payroll_run = get_object_or_404(PayrollRun, pk=pk)
    payslips = payroll_run.payslips.all()

    context = {
        'payroll_run': payroll_run,
        'payslips': payslips,
        'total_employees': payslips.count(),
        'total_net_pay': payslips.aggregate(total=Sum('net_pay'))['total'] or 0,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'url': '/payroll/', 'active': False},
            {'name': payroll_run.run_number, 'active': True},
        ],
    }
    return render(request, 'payroll/run_detail.html', context)


@login_required
@data_entry_or_admin_required
def run_process(request, pk):
    """Process payroll run - Generate payslips"""
    payroll_run = get_object_or_404(PayrollRun, pk=pk)

    if request.method == 'POST':
        # Get all active assignments for this client
        if payroll_run.client:
            assignments = EmployeeAssignment.objects.filter(
                client=payroll_run.client,
                is_current=True
            )
        else:
            assignments = EmployeeAssignment.objects.filter(is_current=True)

        processed_count = 0
        error_count = 0

        for assignment in assignments:
            try:
                # Get attendance data (simplified)
                attendance_data = {
                    'days_present': 26,  # Default - will be fetched from attendance
                    'overtime_hours': 0,
                    'production_units': 0,
                }

                # Calculate salary
                calculator = SalaryCalculator(
                    assignment.employee,
                    assignment,
                    payroll_run.month_year,
                    attendance_data
                )
                result = calculator.process()

                # Create payslip
                payslip = Payslip.objects.create(
                    payroll_run=payroll_run,
                    employee=assignment.employee,
                    assignment=assignment,
                    basic_pay=result['basic_pay'],
                    allowance=result['allowance'],
                    incentive=result['incentive'],
                    conveyance=result['conveyance'],
                    overtime=result['overtime'],
                    other_earnings=result['other_earnings'],
                    gross_earnings=result['gross_earnings'],
                    pf_employee=result['pf_employee'],
                    pf_employer=result['pf_employer'],
                    eps_employer=result['eps_employer'],
                    pf_basic=result['pf_basic'],
                    admin_charges=result['admin_charges'],
                    edlis_charges=result['edlis_charges'],
                    esi_employee=result['esi_employee'],
                    esi_employer=result['esi_employer'],
                    professional_tax=result['professional_tax'],
                    tds=result['tds'],
                    loan_deduction=result['loan_total'],
                    advance_deduction=0,
                    other_deductions=result['other_deductions'],
                    total_deductions=result['total_deductions'],
                    net_pay=result['net_pay'],
                    loan_details=result['loan_deductions'],
                )
                processed_count += 1

            except Exception as e:
                error_count += 1
                print(f"Error processing {assignment.employee.name}: {e}")

        # Update payroll run
        payroll_run.status = 'PROCESSED'
        payroll_run.processed_at = timezone.now()
        payroll_run.total_employees = processed_count
        payroll_run.total_gross_salary = payroll_run.payslips.aggregate(total=Sum('gross_earnings'))['total'] or 0
        payroll_run.total_deductions = payroll_run.payslips.aggregate(total=Sum('total_deductions'))['total'] or 0
        payroll_run.total_net_payable = payroll_run.payslips.aggregate(total=Sum('net_pay'))['total'] or 0
        payroll_run.save()

        messages.success(request, f'Payroll processed! {processed_count} payslips generated.')
        if error_count > 0:
            messages.warning(request, f'{error_count} employees had errors.')

        return redirect('payroll:run_detail', pk=payroll_run.pk)

    context = {
        'payroll_run': payroll_run,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'url': '/payroll/', 'active': False},
            {'name': payroll_run.run_number, 'url': f'/payroll/{payroll_run.id}/', 'active': False},
            {'name': 'Process Payroll', 'active': True},
        ],
    }
    return render(request, 'payroll/run_process.html', context)


@login_required
@data_entry_or_admin_required
def run_finalize(request, pk):
    """Finalize payroll run"""
    payroll_run = get_object_or_404(PayrollRun, pk=pk)

    if request.method == 'POST':
        payroll_run.status = 'FINALIZED'
        payroll_run.finalized_at = timezone.now()
        payroll_run.save()
        messages.success(request, f'Payroll Run {payroll_run.run_number} finalized!')
        return redirect('payroll:run_detail', pk=payroll_run.pk)

    context = {
        'payroll_run': payroll_run,
    }
    return render(request, 'payroll/run_finalize.html', context)


@login_required
@data_entry_or_admin_required
def run_delete(request, pk):
    """Delete payroll run"""
    payroll_run = get_object_or_404(PayrollRun, pk=pk)

    if request.method == 'POST':
        run_number = payroll_run.run_number
        payroll_run.delete()
        messages.success(request, f'Payroll Run {run_number} deleted!')
        return redirect('payroll:run_list')

    context = {
        'payroll_run': payroll_run,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'url': '/payroll/', 'active': False},
            {'name': f'Delete {payroll_run.run_number}', 'active': True},
        ],
    }
    return render(request, 'payroll/run_confirm_delete.html', context)


@login_required
@data_entry_or_admin_required
def run_cancel(request, pk):
    """Cancel payroll run"""
    payroll_run = get_object_or_404(PayrollRun, pk=pk)

    if request.method == 'POST':
        payroll_run.status = 'CANCELLED'
        payroll_run.save()
        messages.success(request, f'Payroll Run {payroll_run.run_number} cancelled!')
        return redirect('payroll:run_list')

    context = {
        'payroll_run': payroll_run,
    }
    return render(request, 'payroll/run_cancel.html', context)


# ============================================
# PAYSLIP VIEWS
# ============================================

@login_required
def payslip_detail(request, pk):
    """View payslip details"""
    payslip = get_object_or_404(Payslip, pk=pk)

    context = {
        'payslip': payslip,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll Runs', 'url': '/payroll/', 'active': False},
            {'name': payslip.payroll_run.run_number, 'url': f'/payroll/{payslip.payroll_run.id}/', 'active': False},
            {'name': payslip.employee.name, 'active': True},
        ],
    }
    return render(request, 'payroll/payslip_detail.html', context)


@login_required
def payslip_download(request, pk):
    """Download payslip as PDF"""
    payslip = get_object_or_404(Payslip, pk=pk)
    # PDF generation logic here
    messages.info(request, 'PDF download coming soon!')
    return redirect('payroll:payslip_detail', pk=payslip.pk)


# ============================================
# PAYMENT VOUCHER VIEWS
# ============================================

@login_required
def payment_voucher_list(request):
    """List payment vouchers"""
    vouchers = PaymentVoucher.objects.all().order_by('-payment_date')

    search = request.GET.get('search')
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(client__name__icontains=search)
        )

    paginator = Paginator(vouchers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'total_vouchers': vouchers.count(),
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payments', 'active': True},
        ],
    }
    return render(request, 'payroll/payment_voucher_list.html', context)


@login_required
@data_entry_or_admin_required
def payment_voucher_create(request):
    """Create payment voucher"""
    if request.method == 'POST':
        form = PaymentVoucherForm(request.POST)
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.created_by = request.user
            voucher.save()
            messages.success(request, f'Payment Voucher {voucher.voucher_number} created!')
            return redirect('payroll:payment_voucher_detail', pk=voucher.pk)
    else:
        form = PaymentVoucherForm()

    context = {
        'form': form,
        'title': 'Create Payment Voucher',
        'button_text': 'Create Payment',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payments', 'url': '/payments/', 'active': False},
            {'name': 'Create Payment', 'active': True},
        ],
    }
    return render(request, 'payroll/payment_voucher_create.html', context)


@login_required
def payment_voucher_detail(request, pk):
    """Payment voucher detail"""
    voucher = get_object_or_404(PaymentVoucher, pk=pk)

    context = {
        'voucher': voucher,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payments', 'url': '/payments/', 'active': False},
            {'name': voucher.voucher_number, 'active': True},
        ],
    }
    return render(request, 'payroll/payment_voucher_detail.html', context)


@login_required
@data_entry_or_admin_required
def payment_voucher_approve(request, pk):
    """Approve payment voucher"""
    voucher = get_object_or_404(PaymentVoucher, pk=pk)

    if request.method == 'POST':
        voucher.status = 'PAID'
        voucher.approved_by = request.user
        voucher.approved_date = timezone.now()
        voucher.paid_at = timezone.now()
        voucher.save()

        # Update payslips
        for detail in voucher.details.all():
            if detail.payslip:
                detail.payslip.paid = True
                detail.payslip.paid_date = timezone.now().date()
                detail.payslip.save()

        messages.success(request, f'Payment Voucher {voucher.voucher_number} approved and paid!')
        return redirect('payroll:payment_voucher_detail', pk=voucher.pk)

    context = {
        'voucher': voucher,
    }
    return render(request, 'payroll/payment_voucher_approve.html', context)


@login_required
@data_entry_or_admin_required
def payment_voucher_cancel(request, pk):
    """Cancel payment voucher"""
    voucher = get_object_or_404(PaymentVoucher, pk=pk)

    if request.method == 'POST':
        voucher.status = 'CANCELLED'
        voucher.save()
        messages.success(request, f'Payment Voucher {voucher.voucher_number} cancelled!')
        return redirect('payroll:payment_voucher_list')

    context = {
        'voucher': voucher,
    }
    return render(request, 'payroll/payment_voucher_cancel.html', context)


# ============================================
# HELPER FUNCTIONS
# ============================================

def get_client_payhead_columns(client):
    """
    Get all payhead columns for a client from ClientPayheadTemplate
    """
    from apps.clients.models import ClientPayheadTemplate

    earnings = []
    deductions = []

    templates = ClientPayheadTemplate.objects.filter(
        client=client,
        is_active=True
    ).order_by('display_order')

    for template in templates:
        if template.type == 'EARNING':
            earnings.append(template.name)
        else:
            deductions.append(template.name)

    return {
        'earnings': earnings,
        'deductions': deductions
    }


def get_client_active_payheads(client, assignments):
    """
    Get payheads that have at least one employee with value > 0
    """
    payhead_columns = get_client_payhead_columns(client)

    active_earnings = []
    active_deductions = []

    # Check each payhead if any assignment has value > 0
    for payhead in payhead_columns['earnings']:
        has_value = False
        for assignment in assignments:
            if assignment.salary_heads.get('earnings', {}).get(payhead, 0) > 0:
                has_value = True
                break
        if has_value:
            active_earnings.append(payhead)

    for payhead in payhead_columns['deductions']:
        has_value = False
        for assignment in assignments:
            if assignment.salary_heads.get('deductions', {}).get(payhead, 0) > 0:
                has_value = True
                break
        if has_value:
            active_deductions.append(payhead)

    return {
        'earnings': active_earnings,
        'deductions': active_deductions
    }


# ============================================
# SALARY SHEET VIEWS
# ============================================

@login_required
def salary_sheet(request):
    """Salary sheet view"""
    clients = Client.objects.filter(is_active=True)
    month_year = request.GET.get('month_year')

    context = {
        'clients': clients,
        'month_year': month_year,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Salary Sheet', 'active': True},
        ],
    }
    return render(request, 'payroll/salary_sheet.html', context)


@login_required
def salary_sheet_generate(request):
    """Generate salary sheet for a client"""
    client_id = request.GET.get('client_id')
    month_year = request.GET.get('month_year')

    if not client_id or not month_year:
        messages.error(request, 'Please select client and month')
        return redirect('payroll:salary_sheet')

    client = get_object_or_404(Client, pk=client_id)

    try:
        month_date = timezone.datetime.strptime(month_year, '%Y-%m-%d').date()
    except ValueError:
        month_date = timezone.now().date().replace(day=1)

    assignments = EmployeeAssignment.objects.filter(
        client=client,
        is_current=True
    )

    payslips = Payslip.objects.filter(
        employee__in=[a.employee for a in assignments],
        payroll_run__month_year=month_date
    )

    context = {
        'client': client,
        'month_year': month_date,
        'assignments': assignments,
        'payslips': payslips,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Salary Sheet', 'url': '/salary-sheet/', 'active': False},
            {'name': f'{client.name} - {month_date.strftime("%B %Y")}', 'active': True},
        ],
    }
    return render(request, 'payroll/salary_sheet_generate.html', context)


@login_required
def salary_sheet_generate_client(request, client_id):
    """Generate salary sheet for a specific client"""
    client = get_object_or_404(Client, pk=client_id)
    month_year = request.GET.get('month_year')

    if not month_year:
        messages.error(request, 'Please select month')
        return redirect('payroll:salary_sheet')

    try:
        month_date = timezone.datetime.strptime(month_year, '%Y-%m-%d').date()
    except ValueError:
        month_date = timezone.now().date().replace(day=1)

    assignments = EmployeeAssignment.objects.filter(
        client=client,
        is_current=True
    )

    active_payheads = get_client_active_payheads(client, assignments)

    data = []
    for assignment in assignments:
        row = {
            'employee': assignment.employee,
            'assignment': assignment,
            'basic_pay': float(assignment.monthly_basic or 0),
        }

        earnings = assignment.salary_heads.get('earnings', {})
        for payhead in active_payheads['earnings']:
            row[payhead] = float(earnings.get(payhead, 0))

        deductions = assignment.salary_heads.get('deductions', {})
        for payhead in active_payheads['deductions']:
            row[payhead] = float(deductions.get(payhead, 0))

        total_earnings = float(assignment.monthly_basic or 0)
        for payhead in active_payheads['earnings']:
            total_earnings += float(earnings.get(payhead, 0))

        total_deductions = 0
        for payhead in active_payheads['deductions']:
            total_deductions += float(deductions.get(payhead, 0))

        row['total_earnings'] = total_earnings
        row['total_deductions'] = total_deductions
        row['net_payable'] = total_earnings - total_deductions

        data.append(row)

    columns = ['S.No', 'Employee Name', 'Employee Code', 'Basic Pay']
    columns.extend(active_payheads['earnings'])
    columns.extend(active_payheads['deductions'])
    columns.extend(['Total Earnings', 'Total Deductions', 'Net Payable'])

    context = {
        'client': client,
        'month_year': month_date,
        'data': data,
        'columns': columns,
        'active_payheads': active_payheads,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Salary Sheet', 'url': '/salary-sheet/', 'active': False},
            {'name': client.name, 'active': True},
        ],
    }
    return render(request, 'payroll/salary_sheet_generate.html', context)


# ============================================
# SALARY SHEET EXPORT
# ============================================

@login_required
def salary_sheet_export(request, client_id):
    """Export salary sheet to Excel"""
    client = get_object_or_404(Client, pk=client_id)
    month_year = request.GET.get('month_year')

    if not month_year:
        messages.error(request, 'Please select month')
        return redirect('payroll:salary_sheet')

    try:
        month_date = timezone.datetime.strptime(month_year, '%Y-%m-%d').date()
    except ValueError:
        month_date = timezone.now().date().replace(day=1)

    assignments = EmployeeAssignment.objects.filter(
        client=client,
        is_current=True
    )

    active_payheads = get_client_active_payheads(client, assignments)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Sheet"

    headers = ['S.No', 'Employee Code', 'Employee Name', 'Basic Pay']
    headers.extend(active_payheads['earnings'])
    headers.extend(active_payheads['deductions'])
    headers.extend(['Total Earnings', 'Total Deductions', 'Net Payable'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, assignment in enumerate(assignments, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=row_idx - 1)
        col += 1
        ws.cell(row=row_idx, column=col, value=assignment.employee.employee_code or '')
        col += 1
        ws.cell(row=row_idx, column=col, value=assignment.employee.name)
        col += 1
        ws.cell(row=row_idx, column=col, value=float(assignment.monthly_basic or 0))
        col += 1

        earnings = assignment.salary_heads.get('earnings', {})
        for payhead in active_payheads['earnings']:
            ws.cell(row=row_idx, column=col, value=float(earnings.get(payhead, 0)))
            col += 1

        deductions = assignment.salary_heads.get('deductions', {})
        for payhead in active_payheads['deductions']:
            ws.cell(row=row_idx, column=col, value=float(deductions.get(payhead, 0)))
            col += 1

        total_earnings = float(assignment.monthly_basic or 0)
        for payhead in active_payheads['earnings']:
            total_earnings += float(earnings.get(payhead, 0))

        total_deductions = 0
        for payhead in active_payheads['deductions']:
            total_deductions += float(deductions.get(payhead, 0))

        ws.cell(row=row_idx, column=col, value=total_earnings)
        col += 1
        ws.cell(row=row_idx, column=col, value=total_deductions)
        col += 1
        ws.cell(row=row_idx, column=col, value=total_earnings - total_deductions)

    for col in range(1, len(headers) + 1):
        max_length = 0
        for row in range(1, min(ws.max_row, 10)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 5, 35)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename=Salary_Sheet_{client.code}_{month_date.strftime("%b_%Y")}.xlsx'
    wb.save(response)
    return response


# apps/payroll/views.py - Add this function

@login_required
def payslip_pdf(request, pk):
    """Download payslip as PDF"""
    payslip = get_object_or_404(Payslip, pk=pk)

    context = {
        'payslip': payslip,
        'basic_pay': payslip.basic_pay,
        'total_earnings': payslip.gross_earnings,
        'total_deductions': payslip.total_deductions,
        'net_payable': payslip.net_pay,
        'earnings': [
            {'name': 'Basic Pay', 'value': payslip.basic_pay},
            {'name': 'Allowance', 'value': payslip.allowance},
            {'name': 'Incentive', 'value': payslip.incentive},
            {'name': 'Conveyance', 'value': payslip.conveyance},
            {'name': 'Overtime', 'value': payslip.overtime},
            {'name': 'Other Earnings', 'value': payslip.other_earnings},
        ],
        'deductions': [
            {'name': 'PF (Employee)', 'value': payslip.pf_employee},
            {'name': 'ESI (Employee)', 'value': payslip.esi_employee},
            {'name': 'Professional Tax', 'value': payslip.professional_tax},
            {'name': 'TDS', 'value': payslip.tds},
            {'name': 'Loan Deduction', 'value': payslip.loan_deduction},
            {'name': 'Advance Deduction', 'value': payslip.advance_deduction},
            {'name': 'Other Deductions', 'value': payslip.other_deductions},
        ],
    }

    # Render HTML template
    return render(request, 'payroll/payslip_pdf.html', context)


# apps/payroll/views.py - Add this view at the end of the file

# ============================================
# EMPLOYEE LEDGER VIEWS
# ============================================

@login_required
def employee_ledger(request, employee_id=None):
    """
    Employee Ledger View
    Shows all transactions for an employee or all employees
    """
    employee = None
    ledgers = EmployeeLedger.objects.all().order_by('-transaction_date')

    if employee_id:
        employee = get_object_or_404(Employee, pk=employee_id)
        ledgers = ledgers.filter(employee=employee)

    # Filter by date range
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date:
        try:
            from_date_obj = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
            ledgers = ledgers.filter(transaction_date__gte=from_date_obj)
        except ValueError:
            pass

    if to_date:
        try:
            to_date_obj = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
            ledgers = ledgers.filter(transaction_date__lte=to_date_obj)
        except ValueError:
            pass

    # Pagination
    paginator = Paginator(ledgers, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Summary
    total_credit = ledgers.filter(transaction_type='CREDIT').aggregate(total=Sum('amount'))['total'] or 0
    total_debit = ledgers.filter(transaction_type='DEBIT').aggregate(total=Sum('amount'))['total'] or 0
    balance = total_credit - total_debit

    context = {
        'employee': employee,
        'page_obj': page_obj,
        'ledgers': ledgers,
        'total_credit': total_credit,
        'total_debit': total_debit,
        'balance': balance,
        'from_date': from_date,
        'to_date': to_date,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Payroll', 'url': '/payroll/', 'active': False},
            {'name': 'Employee Ledger', 'active': True},
        ],
    }
    return render(request, 'payroll/employee_ledger.html', context)


# apps/payroll/views.py - Salary Sheet

@login_required
def generate_salary_sheet(request, client_id, month_year):
    """Generate salary sheet for a client - Dynamic Columns"""
    client = get_object_or_404(Client, pk=client_id)

    assignments = EmployeeAssignment.objects.filter(
        client=client,
        is_current=True
    ).select_related('employee', 'location')

    from apps.payheads.models import Payhead

    # ✅ Get all active fixed payheads
    all_payheads = Payhead.objects.filter(is_active=True).order_by('type', 'display_order', 'name')

    # ✅ Dynamic: Check which payheads have > 0 value for this client
    active_earnings = []
    active_deductions = []

    for payhead in all_payheads:
        has_value = False
        for assignment in assignments:
            if payhead.type in ['EARNING', 'BONUS', 'ALLOWANCE', 'REIMBURSEMENT']:
                if assignment.salary_heads.get('earnings', {}).get(payhead.name, 0) > 0:
                    has_value = True
                    break
            else:
                if assignment.salary_heads.get('deductions', {}).get(payhead.name, 0) > 0:
                    has_value = True
                    break

        if has_value:
            if payhead.type in ['EARNING', 'BONUS', 'ALLOWANCE', 'REIMBURSEMENT']:
                active_earnings.append(payhead.name)
            else:
                active_deductions.append(payhead.name)

    # ✅ Dynamic Columns - Only active payheads
    columns = ['S.No', 'Employee Name', 'Employee Code', 'Basic Pay']
    columns.extend(active_earnings)
    columns.extend(active_deductions)
    columns.extend(['Total Earnings', 'Total Deductions', 'Net Payable'])

    # Build data
    data = []
    for idx, assignment in enumerate(assignments, 1):
        row = {
            's_no': idx,
            'employee_name': assignment.employee.name,
            'employee_code': assignment.employee.employee_code,
            'basic_pay': float(assignment.monthly_basic or 0),
        }

        earnings = assignment.salary_heads.get('earnings', {})
        deductions = assignment.salary_heads.get('deductions', {})

        total_earnings = float(assignment.monthly_basic or 0)
        total_deductions = 0

        for payhead in active_earnings:
            value = float(earnings.get(payhead, 0))
            row[payhead] = value
            total_earnings += value

        for payhead in active_deductions:
            value = float(deductions.get(payhead, 0))
            row[payhead] = value
            total_deductions += value

        row['total_earnings'] = total_earnings
        row['total_deductions'] = total_deductions
        row['net_payable'] = total_earnings - total_deductions

        data.append(row)

    context = {
        'client': client,
        'columns': columns,
        'data': data,
        'active_earnings': active_earnings,
        'active_deductions': active_deductions,
        'total_employees': assignments.count(),
        'month_year': month_year,
    }
    return render(request, 'payroll/salary_sheet.html', context)


# apps/payroll/views.py - Payslip

@login_required
def generate_payslip(request, assignment_id):
    """Generate payslip for an employee - Dynamic Display"""
    assignment = get_object_or_404(EmployeeAssignment, pk=assignment_id)

    from apps.payheads.models import Payhead

    # ✅ Get all active fixed payheads
    all_payheads = Payhead.objects.filter(is_active=True).order_by('type', 'display_order', 'name')

    earnings = {}
    deductions = {}

    for payhead in all_payheads:
        if payhead.type in ['EARNING', 'BONUS', 'ALLOWANCE', 'REIMBURSEMENT']:
            value = assignment.salary_heads.get('earnings', {}).get(payhead.name, 0)
            if value > 0:  # ✅ Dynamic: Only > 0 show
                earnings[payhead.name] = value
        else:
            value = assignment.salary_heads.get('deductions', {}).get(payhead.name, 0)
            if value > 0:  # ✅ Dynamic: Only > 0 show
                deductions[payhead.name] = value

    context = {
        'assignment': assignment,
        'earnings': earnings,
        'deductions': deductions,
        'basic_pay': float(assignment.monthly_basic or 0),
        'total_earnings': assignment.total_earnings,
        'total_deductions': assignment.total_deductions,
        'net_payable': assignment.total_earnings - assignment.total_deductions,
    }
    return render(request, 'payroll/payslip.html', context)