# apps/attendance/views.py
from django.db import models
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.db import transaction
from .models import AttendanceVoucher, AttendanceDetail, AttendanceImportLog
from .forms import AttendanceVoucherForm, AttendanceDetailForm, AttendanceImportForm
from apps.clients.models import Client
from apps.employees.models import Employee
from apps.core.decorators import data_entry_or_admin_required
import openpyxl
from datetime import datetime


@login_required
def attendance_voucher_list(request):
    """List all attendance vouchers with filters"""
    vouchers = AttendanceVoucher.objects.all().order_by('-month_year', '-created_at')

    # Filters
    client_id = request.GET.get('client')
    if client_id:
        vouchers = vouchers.filter(client_id=client_id)

    status = request.GET.get('status')
    if status:
        vouchers = vouchers.filter(status=status)

    month = request.GET.get('month')
    if month:
        try:
            month_date = datetime.strptime(month, '%Y-%m')
            vouchers = vouchers.filter(month_year__year=month_date.year, month_year__month=month_date.month)
        except:
            pass

    search = request.GET.get('search')
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(client__name__icontains=search)
        )

    paginator = Paginator(vouchers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'clients': Client.objects.filter(is_active=True),
        'client_id': client_id,
        'status': status,
        'month': month,
        'search': search,
        'total_vouchers': vouchers.count(),
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_list.html', context)


@login_required
@data_entry_or_admin_required
def attendance_voucher_create(request):
    """Create attendance voucher"""
    if request.method == 'POST':
        form = AttendanceVoucherForm(request.POST)
        if form.is_valid():
            try:
                voucher = form.save(commit=False)
                # ✅ Ensure month_year is first day of month
                if voucher.month_year:
                    voucher.month_year = voucher.month_year.replace(day=1)
                voucher.save()
                messages.success(request, f'Attendance Voucher {voucher.voucher_number} created successfully!')
                return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)
            except Exception as e:
                messages.error(request, f'Error creating voucher: {str(e)}')
    else:
        form = AttendanceVoucherForm()

    context = {
        'form': form,
        'title': 'Create Attendance Voucher',
        'button_text': 'Create Voucher',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': 'Create Voucher', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_form.html', context)

@login_required
def attendance_voucher_detail(request, pk):
    """View attendance voucher with details"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)
    details = voucher.details.all().order_by('employee__name')

    # Search filter for employees
    search = request.GET.get('search')
    if search:
        details = details.filter(
            Q(employee__name__icontains=search) |
            Q(employee__employee_code__icontains=search)
        )

    paginator = Paginator(details, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'voucher': voucher,
        'page_obj': page_obj,
        'details': details,
        'search': search,
        'total_employees': details.count(),
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_detail.html', context)


@login_required
@data_entry_or_admin_required
def attendance_voucher_edit(request, pk):
    """Edit attendance voucher"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        form = AttendanceVoucherForm(request.POST, instance=voucher)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance Voucher updated successfully!')
            return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)
    else:
        form = AttendanceVoucherForm(instance=voucher)

    context = {
        'form': form,
        'voucher': voucher,
        'title': 'Edit Attendance Voucher',
        'button_text': 'Update Voucher',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Edit', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_form.html', context)


@login_required
@data_entry_or_admin_required
def attendance_voucher_delete(request, pk):
    """Delete attendance voucher"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        voucher.delete()
        messages.success(request, 'Attendance Voucher deleted successfully!')
        return redirect('attendance:attendance_voucher_list')

    context = {
        'voucher': voucher,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Delete', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_confirm_delete.html', context)


@login_required
@data_entry_or_admin_required
def attendance_voucher_bulk_delete(request, pk):
    """Bulk delete attendance details"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        detail_ids = request.POST.getlist('detail_ids')

        if not detail_ids:
            messages.warning(request, 'No records selected for deletion.')
            return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

        count = AttendanceDetail.objects.filter(id__in=detail_ids).count()
        AttendanceDetail.objects.filter(id__in=detail_ids).delete()

        # Update voucher totals
        voucher.total_employees = voucher.details.count()
        voucher.total_present_days = voucher.details.aggregate(sum=models.Sum('days_present'))['sum'] or 0
        voucher.total_absent_days = voucher.details.aggregate(sum=models.Sum('days_absent'))['sum'] or 0
        voucher.total_leave_days = voucher.details.aggregate(sum=models.Sum('days_leave'))['sum'] or 0
        voucher.save()

        messages.success(request, f'{count} attendance records deleted successfully!')
        return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

    return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)


@login_required
@data_entry_or_admin_required
def attendance_detail_edit(request, pk):
    """Edit individual attendance detail"""
    detail = get_object_or_404(AttendanceDetail, pk=pk)
    voucher = detail.attendance_voucher

    if request.method == 'POST':
        form = AttendanceDetailForm(request.POST, instance=detail)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance record updated successfully!')
            return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)
    else:
        form = AttendanceDetailForm(instance=detail)

    context = {
        'form': form,
        'detail': detail,
        'voucher': voucher,
        'title': 'Edit Attendance Record',
        'button_text': 'Update Record',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Edit Record', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_detail_form.html', context)


@login_required
@data_entry_or_admin_required
def attendance_detail_delete(request, pk):
    """Delete individual attendance detail"""
    detail = get_object_or_404(AttendanceDetail, pk=pk)
    voucher = detail.attendance_voucher

    if request.method == 'POST':
        detail.delete()

        # Update voucher totals
        voucher.total_employees = voucher.details.count()
        voucher.total_present_days = voucher.details.aggregate(sum=models.Sum('days_present'))['sum'] or 0
        voucher.total_absent_days = voucher.details.aggregate(sum=models.Sum('days_absent'))['sum'] or 0
        voucher.total_leave_days = voucher.details.aggregate(sum=models.Sum('days_leave'))['sum'] or 0
        voucher.save()

        messages.success(request, 'Attendance record deleted successfully!')
        return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

    context = {
        'detail': detail,
        'voucher': voucher,
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Delete Record', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_detail_confirm_delete.html', context)


@login_required
@data_entry_or_admin_required
def attendance_import(request, pk):
    """Import attendance from Excel"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        form = AttendanceImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
                return redirect('attendance:attendance_import', pk=voucher.pk)

            result = import_attendance_from_excel(voucher, excel_file, request.user)

            if result['success']:
                messages.success(request,
                                 f"✅ {result['imported_rows']} records imported successfully, "
                                 f"❌ {result['error_rows']} errors"
                                 )
                if result['errors']:
                    request.session['import_errors'] = result['errors']
                    error_url = reverse('core:export_import_errors')
                    messages.info(request,
                                  f'<a href="{error_url}" class="btn btn-sm btn-danger" target="_blank">'
                                  f'<i class="fas fa-file-excel"></i> 📥 Download Error Report</a>'
                                  )
            else:
                messages.error(request, result.get('message', 'Import failed'))

            return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)
    else:
        form = AttendanceImportForm()

    context = {
        'form': form,
        'voucher': voucher,
        'title': 'Import Attendance',
        'button_text': 'Import',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Import', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_import.html', context)


# apps/attendance/views.py - Updated import_attendance_from_excel

def import_attendance_from_excel(voucher, excel_file, user):
    """Import attendance data from Excel"""
    errors = []
    imported_rows = 0
    error_rows = 0

    # Get total days in month
    total_days = voucher.month_year.day
    if voucher.month_year:
        import calendar
        total_days = calendar.monthrange(voucher.month_year.year, voucher.month_year.month)[1]

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return {
            'success': False,
            'message': f'Error reading file: {str(e)}',
            'errors': [str(e)]
        }

    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip().lower())
        else:
            headers.append('')

    # Find column indices
    emp_code_col = None
    emp_name_col = None
    present_col = None
    leave_col = None

    for idx, header in enumerate(headers):
        if 'employee code' in header or 'emp code' in header:
            emp_code_col = idx
        elif 'employee name' in header or 'emp name' in header or 'name' in header:
            emp_name_col = idx
        elif 'present' in header:
            present_col = idx
        elif 'leave' in header:
            leave_col = idx

    if emp_code_col is None:
        return {
            'success': False,
            'message': 'Employee Code column not found in Excel',
            'errors': ['Employee Code column is required']
        }

    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not any(row):
                continue

            emp_code = str(row[emp_code_col]).strip() if emp_code_col < len(row) and row[emp_code_col] else ''
            if not emp_code:
                errors.append(f"Row {row_idx}: Employee Code is required")
                error_rows += 1
                continue

            employee = Employee.objects.filter(employee_code=emp_code).first()
            if not employee:
                errors.append(f"Row {row_idx}: Employee not found with code: {emp_code}")
                error_rows += 1
                continue

            days_present = float(row[present_col]) if present_col is not None and present_col < len(row) and row[
                present_col] else 0
            days_leave = float(row[leave_col]) if leave_col is not None and leave_col < len(row) and row[
                leave_col] else 0

            # ✅ Auto-calculate absent days
            days_absent = total_days - days_present - days_leave
            if days_absent < 0:
                days_absent = 0
                errors.append(
                    f"Row {row_idx}: Present + Leave ({days_present + days_leave}) exceeds total days ({total_days})")

            detail, created = AttendanceDetail.objects.get_or_create(
                attendance_voucher=voucher,
                employee=employee,
                defaults={
                    'days_present': days_present,
                    'days_absent': days_absent,
                    'days_leave': days_leave,
                }
            )

            if not created:
                detail.days_present = days_present
                detail.days_absent = days_absent
                detail.days_leave = days_leave
                detail.save()

            imported_rows += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            error_rows += 1

    # Update voucher totals
    from django.db import models
    voucher.total_employees = voucher.details.count()
    voucher.total_present_days = voucher.details.aggregate(sum=models.Sum('days_present'))['sum'] or 0
    voucher.total_absent_days = voucher.details.aggregate(sum=models.Sum('days_absent'))['sum'] or 0
    voucher.total_leave_days = voucher.details.aggregate(sum=models.Sum('days_leave'))['sum'] or 0
    voucher.save()

    AttendanceImportLog.objects.create(
        attendance_voucher=voucher,
        file_name=excel_file.name,
        total_rows=ws.max_row - 1,
        imported_rows=imported_rows,
        error_rows=error_rows,
        error_details=errors
    )

    return {
        'success': True,
        'imported_rows': imported_rows,
        'error_rows': error_rows,
        'errors': errors
    }


@login_required
@data_entry_or_admin_required
def attendance_voucher_bulk_delete(request, pk):
    """Bulk delete ALL attendance details for a voucher"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        # Get list of detail IDs
        detail_ids = request.POST.getlist('detail_ids')

        if not detail_ids:
            messages.warning(request, 'No records selected for deletion.')
            return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

        # ✅ Delete all selected records in one query
        count = AttendanceDetail.objects.filter(id__in=detail_ids).count()
        AttendanceDetail.objects.filter(id__in=detail_ids).delete()

        # Update voucher totals
        voucher.total_employees = voucher.details.count()
        voucher.total_present_days = voucher.details.aggregate(sum=models.Sum('days_present'))['sum'] or 0
        voucher.total_absent_days = voucher.details.aggregate(sum=models.Sum('days_absent'))['sum'] or 0
        voucher.total_leave_days = voucher.details.aggregate(sum=models.Sum('days_leave'))['sum'] or 0
        voucher.save()

        messages.success(request, f'{count} attendance records deleted successfully!')
        return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

    return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)


@login_required
@data_entry_or_admin_required
def attendance_voucher_bulk_delete_all(request, pk):
    """Delete ALL attendance details for a voucher (Fast)"""
    voucher = get_object_or_404(AttendanceVoucher, pk=pk)

    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        if confirm and confirm.lower() == 'yes':
            count = voucher.details.count()
            voucher.details.all().delete()

            voucher.total_employees = 0
            voucher.total_present_days = 0
            voucher.total_absent_days = 0
            voucher.total_leave_days = 0
            voucher.save()

            messages.success(request, f'All {count} attendance records deleted successfully!')
        else:
            messages.warning(request, 'Deletion cancelled. Please type "YES" to confirm.')

        return redirect('attendance:attendance_voucher_detail', pk=voucher.pk)

    context = {
        'voucher': voucher,
        'total_count': voucher.details.count(),
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/', 'active': False},
            {'name': 'Attendance', 'url': '/attendance/', 'active': False},
            {'name': voucher.voucher_number, 'url': f'/attendance/{voucher.pk}/', 'active': False},
            {'name': 'Delete All', 'active': True},
        ],
    }
    return render(request, 'attendance/attendance_voucher_bulk_delete_all.html', context)


@login_required
def attendance_download_template(request):
    """Download attendance import template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Import"

    headers = ['Employee Code', 'Employee Name', 'Days Present', 'Days Absent', 'Days Leave']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Sample data
    sample_data = [
        ['EMP0001', 'Rajesh Kumar', 22, 0, 0],
        ['EMP0002', 'Priya Sharma', 20, 1, 1],
        ['EMP0003', 'Amit Singh', 23, 0, 0],
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Attendance_Import_Template.xlsx'
    wb.save(response)
    return response

