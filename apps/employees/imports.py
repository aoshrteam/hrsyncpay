# apps/employees/imports.py
import openpyxl
from decimal import Decimal
from datetime import datetime
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from .models import Employee, EmployeeAssignment
from apps.clients.models import Client


class EmployeeMasterImport:
    """Import Employee Master Data - Auto-generate Master Code"""

    REQUIRED_FIELDS = ['name', 'email', 'phone']

    def __init__(self, excel_file, user):
        self.excel_file = excel_file
        self.user = user
        self.wb = None
        self.ws = None
        self.errors = []
        self.success_count = 0
        self.update_count = 0
        self.error_count = 0

    def process(self):
        """Process the import"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.ws = self.wb.active
        except Exception as e:
            return {
                'success': False,
                'message': f'Error reading Excel file: {str(e)}',
                'errors': [str(e)]
            }

        headers = self.get_headers()

        with transaction.atomic():
            for row_idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    self.process_row(row, headers, row_idx)
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
                    self.error_count += 1

        return {
            'success': True,
            'total_rows': self.ws.max_row - 1,
            'success_count': self.success_count,
            'update_count': self.update_count,
            'error_count': self.error_count,
            'errors': self.errors
        }

    def get_headers(self):
        headers = []
        for cell in self.ws[1]:
            headers.append(cell.value)
        return headers

    def process_row(self, row, headers, row_idx):
        """Process a single row"""
        data = {}
        for idx, header in enumerate(headers):
            if header:
                key = header.lower().replace(' ', '_')
                data[key] = row[idx] if idx < len(row) else None

        # Check required fields
        for field in ['name', 'email', 'phone']:
            if not data.get(field):
                raise Exception(f"Required field '{field}' is missing")

        # Find existing employee (by PAN, Aadhaar, or Email)
        employee = self.find_existing_employee(data)

        if employee:
            self.update_employee(employee, data)
            self.update_count += 1
        else:
            self.create_employee(data)
            self.success_count += 1

    def find_existing_employee(self, data):
        """Find existing employee by PAN, Aadhaar, or Email"""
        pan = data.get('pan_number')
        if pan:
            employee = Employee.objects.filter(pan_number=pan).first()
            if employee:
                return employee

        aadhaar = data.get('aadhaar_number')
        if aadhaar:
            employee = Employee.objects.filter(aadhaar_number=aadhaar).first()
            if employee:
                return employee

        email = data.get('email')
        if email:
            employee = Employee.objects.filter(email=email).first()
            if employee:
                return employee

        return None

    def create_employee(self, data):
        """Create new employee with auto-generated master code"""

        master_code = self.generate_master_code()

        employee = Employee.objects.create(
            master_code=master_code,
            name=data.get('name', ''),
            father_name=data.get('father_name', ''),
            mother_name=data.get('mother_name', ''),
            date_of_birth=self.parse_date(data.get('date_of_birth')),
            gender=data.get('gender', 'M'),
            email=data.get('email', ''),
            phone=data.get('phone', ''),
            alternate_phone=data.get('alternate_phone', ''),
            current_address=data.get('current_address', ''),
            permanent_address=data.get('permanent_address', ''),
            pan_number=data.get('pan_number', ''),
            aadhaar_number=data.get('aadhaar_number', ''),
            pf_number=data.get('pf_number', ''),
            esi_number=data.get('esi_number', ''),
            uan_number=data.get('uan_number', ''),
            bank_name=data.get('bank_name', ''),
            bank_account_number=data.get('bank_account_number', ''),
            ifsc_code=data.get('ifsc_code', ''),
            bank_branch=data.get('bank_branch', ''),
            date_of_joining=self.parse_date(data.get('date_of_joining')),
            is_active=True,
            pf_applicable=data.get('pf_applicable', True),
            pf_employee_rate=Decimal(data.get('pf_employee_rate', 12)),
            pf_employer_rate=Decimal(data.get('pf_employer_rate', 13)),
            esi_applicable=data.get('esi_applicable', True),
            esi_rule=data.get('esi_rule', 'AUTO'),
            esi_limit=Decimal(data.get('esi_limit', 21000)),
            esi_employee_rate=Decimal(data.get('esi_employee_rate', 0.75)),
            esi_employer_rate=Decimal(data.get('esi_employer_rate', 3.25)),
            tds_applicable=data.get('tds_applicable', False),
            tds_type=data.get('tds_type', 'PERCENTAGE'),
            tds_value=Decimal(data.get('tds_value', 0)),
            basic_pay=Decimal(data.get('basic_pay', 0)),
            hra=Decimal(data.get('hra', 0)),
            special_allowance=Decimal(data.get('special_allowance', 0)),
            conveyance=Decimal(data.get('conveyance', 0)),
        )

        return employee

    def update_employee(self, employee, data):
        """Update existing employee"""
        employee.name = data.get('name', employee.name)
        employee.father_name = data.get('father_name', employee.father_name)
        employee.mother_name = data.get('mother_name', employee.mother_name)
        if data.get('date_of_birth'):
            employee.date_of_birth = self.parse_date(data.get('date_of_birth'))
        employee.gender = data.get('gender', employee.gender)
        employee.email = data.get('email', employee.email)
        employee.phone = data.get('phone', employee.phone)
        employee.alternate_phone = data.get('alternate_phone', employee.alternate_phone)
        employee.current_address = data.get('current_address', employee.current_address)
        employee.permanent_address = data.get('permanent_address', employee.permanent_address)
        employee.pan_number = data.get('pan_number', employee.pan_number)
        employee.aadhaar_number = data.get('aadhaar_number', employee.aadhaar_number)
        employee.pf_number = data.get('pf_number', employee.pf_number)
        employee.esi_number = data.get('esi_number', employee.esi_number)
        employee.uan_number = data.get('uan_number', employee.uan_number)
        employee.bank_name = data.get('bank_name', employee.bank_name)
        employee.bank_account_number = data.get('bank_account_number', employee.bank_account_number)
        employee.ifsc_code = data.get('ifsc_code', employee.ifsc_code)
        employee.bank_branch = data.get('bank_branch', employee.bank_branch)
        if data.get('date_of_joining'):
            employee.date_of_joining = self.parse_date(data.get('date_of_joining'))
        employee.save()

    def generate_master_code(self):
        """Generate auto master code: EMP0001, EMP0002..."""
        last_employee = Employee.objects.all().order_by('-id').first()
        if last_employee and last_employee.master_code:
            try:
                code = last_employee.master_code
                num_str = ''.join(filter(str.isdigit, code))
                if num_str:
                    last_num = int(num_str)
                    new_num = last_num + 1
                else:
                    new_num = 1
            except (ValueError, IndexError):
                new_num = 1
        else:
            new_num = 1

        return f"EMP{new_num:04d}"

    def parse_date(self, value):
        """Parse date from various formats"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                try:
                    return datetime.strptime(value, '%d-%m-%Y').date()
                except ValueError:
                    return None
        return None

    def get_sample_template(self):
        """Generate sample Excel template for Employee Master Import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Master"

        headers = [
            # Personal Details
            'Name', 'Father Name', 'Mother Name', 'Date of Birth', 'Gender',
            'Email', 'Phone', 'Alternate Phone', 'Current Address', 'Permanent Address',

            # Statutory IDs
            'PAN Number', 'Aadhaar Number', 'PF Number', 'ESI Number', 'UAN Number',

            # Bank Details
            'Bank Name', 'Bank Account Number', 'IFSC Code', 'Bank Branch',

            # Employment
            'Date of Joining', 'PF Applicable', 'PF Employee Rate', 'PF Employer Rate',
            'ESI Applicable', 'ESI Rule', 'ESI Limit', 'ESI Employee Rate', 'ESI Employer Rate',
            'TDS Applicable', 'TDS Type', 'TDS Value',
            'Basic Pay', 'HRA', 'Special Allowance', 'Conveyance'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        # Sample data
        sample_data = [
            ['Rajesh Kumar', 'Ramesh Kumar', 'Sita Devi', '1990-01-15', 'M',
             'rajesh@email.com', '9876543210', '', '123, Jaipur', '',
             'ABCDE1234F', '123456789012', 'PF12345', 'ESI12345', 'UAN12345',
             'HDFC Bank', '1234567890', 'HDFC0001', 'Jaipur Branch',
             '2024-01-01', 'Yes', '12', '13',
             'Yes', 'AUTO', '21000', '0.75', '3.25',
             'No', 'PERCENTAGE', '0',
             '25000', '10000', '5000', '2000']
        ]

        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, 4):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2

        return wb


class AssignmentImport:
    """Import Employee Assignments - Requires Master Code + Client Code"""

    REQUIRED_FIELDS = ['master_code', 'client_code', 'start_date']

    def __init__(self, excel_file, user):
        self.excel_file = excel_file
        self.user = user
        self.wb = None
        self.ws = None
        self.errors = []
        self.success_count = 0
        self.error_count = 0

    def process(self):
        """Process assignment import"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.ws = self.wb.active
        except Exception as e:
            return {
                'success': False,
                'message': f'Error reading Excel file: {str(e)}',
                'errors': [str(e)]
            }

        headers = self.get_headers()

        with transaction.atomic():
            for row_idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    self.process_row(row, headers, row_idx)
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
                    self.error_count += 1

        return {
            'success': True,
            'total_rows': self.ws.max_row - 1,
            'success_count': self.success_count,
            'error_count': self.error_count,
            'errors': self.errors
        }

    def get_headers(self):
        headers = []
        for cell in self.ws[1]:
            headers.append(cell.value)
        return headers

    def process_row(self, row, headers, row_idx):
        """Process a single assignment row"""
        data = {}
        for idx, header in enumerate(headers):
            if header:
                key = header.lower().replace(' ', '_')
                data[key] = row[idx] if idx < len(row) else None

        # Check required fields
        for field in ['master_code', 'client_code', 'start_date']:
            if not data.get(field):
                raise Exception(f"Required field '{field}' is missing")

        # Find employee by master code
        master_code = data.get('master_code')
        employee = Employee.objects.filter(master_code=master_code).first()
        if not employee:
            raise Exception(f"Employee with master code '{master_code}' not found")

        # Find client by client code
        client_code = data.get('client_code')
        client = Client.objects.filter(code=client_code).first()
        if not client:
            raise Exception(f"Client with code '{client_code}' not found")

        # Parse dates
        start_date = self.parse_date(data.get('start_date'))
        end_date = self.parse_date(data.get('end_date'))
        pf_joining = self.parse_date(data.get('pf_joining_date')) or start_date
        pf_exit = self.parse_date(data.get('pf_exit_date'))

        if not start_date:
            raise Exception("Start date is required")

        # Check if assignment exists
        assignment = EmployeeAssignment.objects.filter(
            employee=employee,
            client=client
        ).filter(
            # Check if assignment exists with same start date or overlapping
            start_date__lte=end_date if end_date else timezone.now().date(),
            end_date__gte=start_date
        ).first()

        if assignment:
            # Update existing assignment
            assignment.end_date = end_date
            assignment.is_current = end_date is None
            assignment.pf_joining_date = pf_joining
            assignment.pf_exit_date = pf_exit
            assignment.save()
        else:
            # Generate assignment code
            assignment_code = self.generate_assignment_code(employee, client)

            # Create new assignment
            assignment = EmployeeAssignment.objects.create(
                employee=employee,
                client=client,
                assignment_code=assignment_code,
                start_date=start_date,
                end_date=end_date,
                is_current=end_date is None,
                pf_joining_date=pf_joining,
                pf_exit_date=pf_exit,
                status='ACTIVE' if end_date is None else 'EXIT'
            )

        self.success_count += 1

    def generate_assignment_code(self, employee, client):
        """Generate assignment code: ClientCode-MasterCode-Sequence"""
        client_code = client.code
        master_code = employee.master_code

        # Get last assignment for this employee and client
        last_assignment = EmployeeAssignment.objects.filter(
            employee=employee,
            client=client
        ).order_by('-id').first()

        if last_assignment and last_assignment.assignment_code:
            parts = last_assignment.assignment_code.split('-')
            if len(parts) == 3:
                try:
                    last_sequence = int(parts[2])
                    new_sequence = last_sequence + 1
                except ValueError:
                    new_sequence = 1
            else:
                new_sequence = 1
        else:
            new_sequence = 1

        return f"{client_code}-{master_code}-{new_sequence:03d}"

    def parse_date(self, value):
        """Parse date from various formats"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                try:
                    return datetime.strptime(value, '%d-%m-%Y').date()
                except ValueError:
                    return None
        return None

    def get_sample_template(self):
        """Generate sample Excel template for Assignment Import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assignment Import"

        headers = [
            'Master Code (Required)', 'Client Code (Required)',
            'Start Date (Required)', 'End Date',
            'PF Joining Date', 'PF Exit Date',
            'Salary Method', 'Monthly Basic'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        # Sample data
        sample_data = [
            ['EMP0001', 'ABC', '2024-01-01', '2024-06-30', '2024-01-01', '2024-06-30', 'CALENDAR_MONTH', '25000'],
            ['EMP0002', 'IHC', '2024-02-01', '', '2024-02-01', '', 'PER_DAY', '2000'],
            ['EMP0003', 'XYZ', '2024-03-01', '2024-08-31', '2024-03-01', '2024-08-31', '26_DAYS_MONTH', '30000'],
        ]

        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, 4):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2

        return wb