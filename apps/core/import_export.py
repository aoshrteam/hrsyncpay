# apps/core/import_export.py - Complete Fixed Version

import openpyxl
import json
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from apps.employees.models import Employee, EmployeeAssignment
from apps.clients.models import Client
from apps.company.models import Company
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re


class ImportExportService:
    """Centralized Import/Export Service"""

    # ============================================
    # HELPER METHODS (unchanged)
    # ============================================

    @staticmethod
    def _get_value(cell):
        if cell is None:
            return None
        if hasattr(cell, 'value'):
            value = cell.value
        else:
            value = cell
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def _safe_str(val):
        if val is None:
            return ''
        return str(val).strip()

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            formats = [
                '%d-%b-%y', '%d-%b-%Y', '%d/%m/%Y', '%Y-%m-%d',
                '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%y', '%d %b %Y'
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
            try:
                if value.isdigit():
                    return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(value) - 2).date()
            except:
                pass
        return None

    @staticmethod
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
        if hasattr(value, 'date'):
            return value.date()
        date_str = str(value).strip()
        date_str = date_str.replace('.', '-').replace('/', '-')
        date_str = re.sub(r'\s+', ' ', date_str)

        formats = [
            ('%d-%b-%Y', True), ('%d-%B-%Y', True),
            ('%d-%b-%y', True), ('%d-%B-%y', True),
            ('%d %b %Y', True), ('%d %B %Y', True),
            ('%d-%m-%Y', False), ('%d/%m/%Y', False),
            ('%d.%m.%Y', False), ('%d-%m-%y', False),
            ('%Y-%m-%d', False),
        ]

        for fmt, case_sensitive in formats:
            try:
                if case_sensitive:
                    try:
                        parsed = datetime.strptime(date_str, fmt)
                        return parsed.date()
                    except ValueError:
                        pass
                    try:
                        parsed = datetime.strptime(date_str.title(), fmt)
                        return parsed.date()
                    except ValueError:
                        pass
                    try:
                        parsed = datetime.strptime(date_str.upper(), fmt)
                        return parsed.date()
                    except ValueError:
                        pass
                else:
                    parsed = datetime.strptime(date_str, fmt)
                    return parsed.date()
            except ValueError:
                continue

        try:
            if date_str.isdigit():
                from datetime import timedelta
                base = datetime(1900, 1, 1)
                parsed = base + timedelta(days=int(date_str) - 2)
                return parsed.date()
        except:
            pass
        return None

    @staticmethod
    def _generate_master_code():
        last = Employee.objects.all().order_by('-id').first()
        if last and last.master_code:
            try:
                num = int(''.join(filter(str.isdigit, last.master_code)))
                return f"EMP{num + 1:04d}"
            except:
                pass
        return "EMP0001"

    @staticmethod
    def _generate_employee_code():
        last = Employee.objects.all().order_by('-id').first()
        if last and last.employee_code:
            try:
                num = int(''.join(filter(str.isdigit, last.employee_code)))
                return f"EMP{num + 1:04d}"
            except:
                pass
        return "EMP0001"

    @staticmethod
    def _get_or_none(model, **kwargs):
        try:
            return model.objects.get(**kwargs)
        except model.DoesNotExist:
            return None

    @staticmethod
    def _get_nullable_value(value):
        val = ImportExportService._safe_str(value)
        return None if val == '' else val

    @staticmethod
    def _parse_decimal(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            try:
                return Decimal(value)
            except:
                return None
        return None

    @staticmethod
    def _find_employee(master_code=None, employee_code=None):
        if master_code:
            employee = ImportExportService._get_or_none(Employee, master_code=master_code)
            if employee:
                return employee
        if employee_code:
            employee = ImportExportService._get_or_none(Employee, employee_code=employee_code)
            if employee:
                return employee
        return None

    # ============================================
    # PAYHEAD TYPE MAPPING
    # ============================================

    PAYHEAD_TYPE_MAPPING = {
        'bonus': 'BONUS', 'da': 'EARNING', 'hra': 'ALLOWANCE',
        'conveyance': 'ALLOWANCE', 'incentive': 'EARNING',
        'special_allowance': 'ALLOWANCE', 'overtime': 'EARNING',
        'medical_allowance': 'ALLOWANCE', 'lta': 'ALLOWANCE',
        'travel_allowance': 'ALLOWANCE', 'phone_allowance': 'ALLOWANCE',
        'internet_allowance': 'ALLOWANCE', 'night_allowance': 'ALLOWANCE',
        'shift_allowance': 'ALLOWANCE', 'performance_bonus': 'BONUS',
        'annual_bonus': 'BONUS', 'diwali_bonus': 'BONUS', 'festival_bonus': 'BONUS',
        'insurance': 'DEDUCTION', 'advance': 'DEDUCTION',
        'professional_tax': 'STATUTORY_DEDUCTION', 'pt': 'STATUTORY_DEDUCTION',
        'loan_recovery': 'LOAN', 'loan': 'LOAN', 'recovery': 'LOAN',
        'tds': 'STATUTORY_DEDUCTION', 'income_tax': 'STATUTORY_DEDUCTION',
        'gratuity': 'DEDUCTION', 'union_fee': 'DEDUCTION',
        'welfare_fund': 'DEDUCTION', 'other_deduction': 'DEDUCTION',
        'canteen': 'DEDUCTION', 'accommodation': 'DEDUCTION', 'transport': 'DEDUCTION',
        'pf': 'STATUTORY_DEDUCTION', 'esi': 'STATUTORY_DEDUCTION',
        'epf': 'STATUTORY_DEDUCTION', 'eps': 'STATUTORY_DEDUCTION',
        'pension': 'STATUTORY_DEDUCTION', 'provident_fund': 'STATUTORY_DEDUCTION',
        'employee_pf': 'STATUTORY_DEDUCTION', 'employer_pf': 'STATUTORY_DEDUCTION',
        'medical_reimbursement': 'REIMBURSEMENT', 'travel_reimbursement': 'REIMBURSEMENT',
        'phone_reimbursement': 'REIMBURSEMENT', 'internet_reimbursement': 'REIMBURSEMENT',
        'fuel_reimbursement': 'REIMBURSEMENT',
    }

    @staticmethod
    def get_payhead_type(payhead_name):
        name_lower = payhead_name.lower().strip().replace(' ', '_')
        if name_lower in ImportExportService.PAYHEAD_TYPE_MAPPING:
            return ImportExportService.PAYHEAD_TYPE_MAPPING[name_lower]
        for key, value in ImportExportService.PAYHEAD_TYPE_MAPPING.items():
            if key in name_lower or name_lower in key:
                return value
        return 'EARNING'

    # ============================================
    # EMPLOYEE IMPORT - COMPLETE
    # ============================================

    @staticmethod
    def import_employees(excel_file, user):
        errors = []
        success_count = 0
        update_count = 0
        create_count = 0
        error_count = 0

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            return {
                'success': False,
                'message': f'Error reading file: {str(e)}',
                'errors': [str(e)]
            }

        headers = []
        data_start_row = 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_values = [str(v).strip() if v else '' for v in row]
            if any('Employee Number' in v or 'S.No' in v or 'Employee Code' in v for v in row_values):
                headers = row_values
                data_start_row = row_idx + 1
                break

        if not headers or all(h == '' for h in headers):
            headers = [str(cell).strip() if cell else '' for cell in ws[2]]
            data_start_row = 3

        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                h_clean = h.lower().strip().replace(' ', '_').replace('/', '_')
                header_map[h_clean] = idx

        for row_idx in range(data_start_row, ws.max_row + 1):
            try:
                row = ws[row_idx]
                row_data = {}

                for key, col_idx in header_map.items():
                    if col_idx < len(row):
                        row_data[key] = ImportExportService._get_value(row[col_idx])

                if not any(row_data.values()):
                    continue

                master_code = ImportExportService._get_nullable_value(row_data.get('employee_number', ''))
                employee_code = ImportExportService._get_nullable_value(row_data.get('employee_code', ''))

                if not master_code and not employee_code:
                    errors.append({
                        'row': row_idx,
                        'column': 'Employee Number / Employee Code',
                        'error': 'Either Master Code or Employee Code is required',
                        'value': ''
                    })
                    error_count += 1
                    continue

                employee = ImportExportService._find_employee(
                    master_code=master_code,
                    employee_code=employee_code
                )

                if employee:
                    ImportExportService._update_employee_smart(employee, row_data)
                    update_count += 1
                else:
                    ImportExportService._create_employee_from_excel(row_data)
                    create_count += 1

                success_count += 1

            except Exception as e:
                errors.append({
                    'row': row_idx,
                    'column': 'General',
                    'error': str(e),
                    'value': ''
                })
                error_count += 1

        return {
            'success': True,
            'total_rows': ws.max_row - data_start_row + 1,
            'create_count': create_count,
            'update_count': update_count,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors,
        }

    # ============================================
    # ✅ FIXED: CREATE EMPLOYEE FROM EXCEL
    # ============================================

    @staticmethod
    def _create_employee_from_excel(data):
        master_code = ImportExportService._get_nullable_value(data.get('employee_number', ''))
        employee_code = ImportExportService._get_nullable_value(data.get('employee_code', ''))

        if not master_code:
            master_code = ImportExportService._generate_master_code()

        if not employee_code:
            employee_code = ImportExportService._generate_employee_code()
        elif Employee.objects.filter(employee_code=employee_code).exists():
            employee_code = ImportExportService._generate_employee_code()

        date_of_birth = ImportExportService._parse_date(data.get('date_of_birth', ''))
        joining_date = ImportExportService._parse_date(data.get('joining_date', ''))

        gender_map = {'MALE': 'M', 'FEMALE': 'F', 'OTHER': 'O', 'M': 'M', 'F': 'F', 'O': 'O'}
        gender = data.get('gender', 'M')
        if gender:
            gender = str(gender).upper().strip()
            gender = gender_map.get(gender, 'M')

        employee = Employee.objects.create(
            master_code=master_code,
            employee_code=employee_code,
            name=ImportExportService._safe_str(data.get('name_of_the_employee', '')),
            father_name=ImportExportService._safe_str(data.get('father_/_mother_name', '')),
            mother_name=ImportExportService._safe_str(data.get('spouse_name', '')),
            alternate_phone=ImportExportService._safe_str(data.get('alternate_phone', '')),
            current_address=ImportExportService._safe_str(data.get('current_address', '')),
            permanent_address=ImportExportService._safe_str(data.get('permanent_address', '')),
            bank_name=ImportExportService._safe_str(data.get('bank_name', '')),
            ifsc_code=ImportExportService._safe_str(data.get('ifs_code', '')),
            bank_branch=ImportExportService._safe_str(data.get('branch', '')),
            date_of_birth=date_of_birth,
            gender=gender,
            email=ImportExportService._get_nullable_value(data.get('email', '')),
            phone=ImportExportService._get_nullable_value(data.get('phone', '')),
            pan_number=ImportExportService._get_nullable_value(data.get('income_tax_number_pan', '')),
            aadhaar_number=ImportExportService._get_nullable_value(data.get('aadhaar_number', '')),
            pf_number=ImportExportService._get_nullable_value(data.get('pf_account_number', '')),
            esi_number=ImportExportService._get_nullable_value(data.get('esi_number', '')),
            uan_number=ImportExportService._get_nullable_value(data.get('universal_account_number_uan', '')),
            bank_account_number=ImportExportService._get_nullable_value(data.get('account_number', '')),
            date_of_joining=joining_date,
            is_active=True,
            pf_applicable=bool(data.get('pf_account_number', '')),
            pf_employee_rate=12.00,
            pf_employer_rate=13.00,
            pf_capping='CAPPED_15000',
            esi_applicable=bool(data.get('esi_number', '')),
            esi_rule='AUTO',
            esi_limit=21000.00,
            esi_employee_rate=0.75,
            esi_employer_rate=3.25,
        )

        return employee

    # ============================================
    # ✅ FIXED: UPDATE EMPLOYEE - ALL FIELDS
    # ============================================

    @staticmethod
    # apps/core/import_export.py - _update_employee_smart

    @staticmethod
    def _update_employee_smart(employee, data):
        """Update employee - DON'T swap codes"""

        # ✅ Update master_code only if provided
        master_code = ImportExportService._get_nullable_value(data.get('employee_number', ''))
        if master_code and master_code != employee.master_code:
            if not Employee.objects.filter(master_code=master_code).exclude(pk=employee.pk).exists():
                employee.master_code = master_code

        # ✅ Update employee_code only if provided
        employee_code = ImportExportService._get_nullable_value(data.get('employee_code', ''))
        if employee_code and employee_code != employee.employee_code:
            if not Employee.objects.filter(employee_code=employee_code).exclude(pk=employee.pk).exists():
                employee.employee_code = employee_code

        # ... rest of update

        name = ImportExportService._safe_str(data.get('name_of_the_employee', ''))
        if name:
            employee.name = name
            updated = True

        father_name = ImportExportService._safe_str(data.get('father_/_mother_name', ''))
        if father_name:
            employee.father_name = father_name
            updated = True

        mother_name = ImportExportService._safe_str(data.get('spouse_name', ''))
        if mother_name:
            employee.mother_name = mother_name
            updated = True

        dob = ImportExportService._parse_date(data.get('date_of_birth', ''))
        if dob:
            employee.date_of_birth = dob
            updated = True

        gender = data.get('gender', '')
        if gender:
            gender_map = {'MALE': 'M', 'FEMALE': 'F', 'OTHER': 'O', 'M': 'M', 'F': 'F', 'O': 'O'}
            gender = str(gender).upper().strip()
            employee.gender = gender_map.get(gender, 'M')
            updated = True

        email = ImportExportService._get_nullable_value(data.get('email', ''))
        if email is not None:
            if not Employee.objects.filter(email=email).exclude(pk=employee.pk).exists():
                employee.email = email
                updated = True

        phone = ImportExportService._get_nullable_value(data.get('phone', ''))
        if phone is not None:
            employee.phone = phone
            updated = True

        alternate_phone = ImportExportService._safe_str(data.get('alternate_phone', ''))
        if alternate_phone:
            employee.alternate_phone = alternate_phone
            updated = True

        current_address = ImportExportService._safe_str(data.get('current_address', ''))
        if current_address:
            employee.current_address = current_address
            updated = True

        permanent_address = ImportExportService._safe_str(data.get('permanent_address', ''))
        if permanent_address:
            employee.permanent_address = permanent_address
            updated = True

        pan = ImportExportService._get_nullable_value(data.get('income_tax_number_pan', ''))
        if pan is not None:
            if not Employee.objects.filter(pan_number=pan).exclude(pk=employee.pk).exists():
                employee.pan_number = pan
                updated = True

        aadhaar = ImportExportService._get_nullable_value(data.get('aadhaar_number', ''))
        if aadhaar is not None:
            if not Employee.objects.filter(aadhaar_number=aadhaar).exclude(pk=employee.pk).exists():
                employee.aadhaar_number = aadhaar
                updated = True

        pf_number = ImportExportService._get_nullable_value(data.get('pf_account_number', ''))
        if pf_number is not None:
            if not Employee.objects.filter(pf_number=pf_number).exclude(pk=employee.pk).exists():
                employee.pf_number = pf_number
                updated = True

        esi_number = ImportExportService._get_nullable_value(data.get('esi_number', ''))
        if esi_number is not None:
            if not Employee.objects.filter(esi_number=esi_number).exclude(pk=employee.pk).exists():
                employee.esi_number = esi_number
                updated = True

        uan = ImportExportService._get_nullable_value(data.get('universal_account_number_uan', ''))
        if uan is not None:
            employee.uan_number = uan
            updated = True

        bank_name = ImportExportService._safe_str(data.get('bank_name', ''))
        if bank_name:
            employee.bank_name = bank_name
            updated = True

        account_number = ImportExportService._get_nullable_value(data.get('account_number', ''))
        if account_number is not None:
            if not Employee.objects.filter(bank_account_number=account_number).exclude(pk=employee.pk).exists():
                employee.bank_account_number = account_number
                updated = True

        ifsc = ImportExportService._safe_str(data.get('ifs_code', ''))
        if ifsc:
            employee.ifsc_code = ifsc
            updated = True

        branch = ImportExportService._safe_str(data.get('branch', ''))
        if branch:
            employee.bank_branch = branch
            updated = True

        joining_date = ImportExportService._parse_date(data.get('joining_date', ''))
        if joining_date:
            employee.date_of_joining = joining_date
            updated = True

        if pf_number:
            employee.pf_applicable = True
            updated = True

        if esi_number:
            employee.esi_applicable = True
            updated = True

        if updated:
            employee.save()

        return employee

    # ============================================
    # ✅ FIXED: ASSIGNMENT IMPORT - ONLY ASSIGNMENT
    # ============================================

    # apps/core/import_export.py - Complete Fixed import_assignments

    # apps/core/import_export.py - Fix import_assignments

    @staticmethod
    def import_assignments(excel_file, user):
        """
        Import Employee Assignments from Excel
        ✅ Find employee by Employee Code OR Master Code
        ✅ Employee Master Data is NEVER modified
        """
        errors = []
        success_count = 0
        update_count = 0
        create_count = 0
        error_count = 0
        created_payheads = []
        updated_payheads = []

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            return {
                'success': False,
                'message': f'Error reading file: {str(e)}',
                'errors': [str(e)]
            }

        # Get headers
        headers = []
        data_start_row = 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            row_values = [str(v).strip() if v else '' for v in row]
            if any('Employee Code' in v or 'Client Code' in v for v in row_values):
                headers = row_values
                data_start_row = row_idx + 1
                break

        if not headers:
            headers = [str(cell).strip() if cell else '' for cell in ws[1]]
            data_start_row = 2

        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                h_clean = h.lower().strip().replace(' ', '_').replace('(', '').replace(')', '').replace('*', '')
                header_map[h_clean] = idx

        # ============================================
        # IDENTIFY PAYHEAD COLUMNS
        # ============================================
        from apps.payheads.models import Payhead

        BASIC_COLUMNS = [
            'employee_code', 'master_code', 'employee_name',
            'client_code', 'location_code',
            'start_date', 'end_date', 'effective_date', 'salary_method',
            'monthly_basic', 'per_day_rate', 'rate_per_unit', 'pf_cap',
            'esi_rule', 'eps_applicable', 'eps_employer_rate', 'eps_limit',
            'professional_tax_exempt', 'other_deductions', 'status'
        ]

        payhead_columns = {}
        for key, idx in header_map.items():
            if key not in BASIC_COLUMNS:
                payhead_columns[key] = idx

        # ============================================
        # AUTO-CREATE PAYHEADS
        # ============================================
        for payhead_name in payhead_columns.keys():
            clean_name = payhead_name.title().replace('_', ' ')
            payhead_type = ImportExportService.get_payhead_type(payhead_name)

            payhead, created = Payhead.objects.get_or_create(
                name=clean_name,
                defaults={
                    'type': payhead_type,
                    'is_active': True,
                    'code': payhead_name.upper().replace(' ', '_'),
                }
            )

            if created:
                created_payheads.append(f"{clean_name} ({payhead_type})")
            elif payhead.type != payhead_type:
                old_type = payhead.type
                payhead.type = payhead_type
                payhead.save()
                updated_payheads.append(f"{clean_name}: {old_type} → {payhead_type}")

        # ============================================
        # DEFAULTS
        # ============================================
        DEFAULTS = {
            'salary_method': 'CALENDAR_MONTH',
            'pf_cap': 'CAPPED_15000',
            'esi_rule': 'AUTO',
            'eps_applicable': 'YES',
            'eps_employer_rate': '8.33',
            'eps_limit': '15000',
            'professional_tax_exempt': 'NO',
            'status': 'ACTIVE'
        }

        VALID_VALUES = {
            'salary_method': ['CALENDAR_MONTH', '26_DAYS_MONTH', 'PER_DAY', 'PRODUCTION'],
            'pf_cap': ['CAPPED_15000', 'CAPPED_18000', 'FULL', 'NOT_APPLICABLE'],
            'esi_rule': ['AUTO', 'FORCE', 'EXEMPT'],
            'status': ['ACTIVE', 'COMPLETED', 'TERMINATED', 'TRANSFERRED']
        }

        # ============================================
        # PROCESS EACH ROW
        # ============================================
        for row_idx in range(data_start_row, ws.max_row + 1):
            try:
                row = ws[row_idx]
                row_data = {}

                for key, col_idx in header_map.items():
                    if col_idx < len(row):
                        row_data[key] = ImportExportService._get_value(row[col_idx])

                    if not any(row_data.values()):
                        continue

                # ============================================
                # GET FIELDS
                # ============================================
                employee_code = ImportExportService._safe_str(row_data.get('employee_code', ''))
                master_code = ImportExportService._safe_str(row_data.get('master_code', ''))
                employee_name = ImportExportService._safe_str(row_data.get('employee_name', ''))
                client_code = ImportExportService._safe_str(row_data.get('client_code', ''))
                location_code = ImportExportService._safe_str(row_data.get('location_code', ''))

                # ============================================
                # ✅ VALIDATE: Employee Code OR Master Code required
                # ============================================
                if not employee_code and not master_code:
                    errors.append({
                        'row': row_idx,
                        'column': 'Employee Code / Master Code',
                        'error': 'Either Employee Code or Master Code is required',
                        'value': ''
                    })
                    error_count += 1
                    continue

                if not client_code:
                    errors.append({
                        'row': row_idx,
                        'column': 'Client Code',
                        'error': 'Client Code is required',
                        'value': ''
                    })
                    error_count += 1
                    continue

                # ============================================
                # ✅ FIND EMPLOYEE - TRY BOTH CODES
                # ============================================
                employee = None

                # ✅ First try: Find by Employee Code
                if employee_code:
                    # Remove any extra text like "(Employee Code (use with Master Code))"
                    clean_code = employee_code.split('(')[0].strip()
                    employee = ImportExportService._get_or_none(Employee, employee_code=clean_code)

                # ✅ Second try: Find by Master Code
                if not employee and master_code:
                    employee = ImportExportService._get_or_none(Employee, master_code=master_code)

                # ✅ Third try: If employee_code is actually master_code
                if not employee and employee_code:
                    # Try finding by master_code using employee_code value
                    employee = ImportExportService._get_or_none(Employee, master_code=employee_code)

                if not employee:
                    errors.append({
                        'row': row_idx,
                        'column': 'Employee Code / Master Code',
                        'error': f'Employee not found. Checked Code: {employee_code}, Master: {master_code}',
                        'value': employee_code or master_code
                    })
                    error_count += 1
                    continue

                # ============================================
                # ✅ VALIDATE EMPLOYEE NAME (Warning only)
                # ============================================
                if employee_name:
                    db_name = employee.name.strip().lower() if employee.name else ''
                    excel_name = employee_name.strip().lower()

                    if db_name != excel_name:
                        errors.append({
                            'row': row_idx,
                            'column': 'Employee Name',
                            'error': f'Name mismatch! Database: "{employee.name}", Excel: "{employee_name}"',
                            'value': employee_name
                        })
                        error_count += 1
                        # Continue anyway

                # ============================================
                # FIND CLIENT
                # ============================================
                client = ImportExportService._get_or_none(Client, code=client_code)
                if not client:
                    errors.append({
                        'row': row_idx,
                        'column': 'Client Code',
                        'error': f'Client not found with Code: {client_code}',
                        'value': client_code
                    })
                    error_count += 1
                    continue

                # ============================================
                # FIND LOCATION
                # ============================================
                location = None
                if location_code:
                    from apps.locations.models import Location
                    location = ImportExportService._get_or_none(Location, location_code=location_code, client=client)
                    if not location:
                        errors.append({
                            'row': row_idx,
                            'column': 'Location Code',
                            'error': f'Location not found with Code: {location_code} for client: {client.name}',
                            'value': location_code
                        })
                        error_count += 1
                        continue
                else:
                    errors.append({
                        'row': row_idx,
                        'column': 'Location Code',
                        'error': 'Location Code is required',
                        'value': ''
                    })
                    error_count += 1
                    continue

                # ============================================
                # PARSE DATES
                # ============================================
                start_date = ImportExportService.parse_date_flexible(row_data.get('start_date', ''))
                end_date = ImportExportService.parse_date_flexible(row_data.get('end_date', ''))
                effective_date = ImportExportService.parse_date_flexible(row_data.get('effective_date', ''))

                if not start_date:
                    errors.append({
                        'row': row_idx,
                        'column': 'Start Date',
                        'error': 'Start Date is required',
                        'value': row_data.get('start_date', '')
                    })
                    error_count += 1
                    continue

                if not effective_date:
                    effective_date = start_date

                # ============================================
                # DROPDOWN VALUES
                # ============================================
                salary_method = ImportExportService._safe_str(row_data.get('salary_method', ''))
                if salary_method and salary_method.upper() in [v.upper() for v in VALID_VALUES['salary_method']]:
                    salary_method = salary_method.upper()
                else:
                    salary_method = DEFAULTS['salary_method']

                pf_cap = ImportExportService._safe_str(row_data.get('pf_cap', ''))
                if pf_cap and pf_cap.upper() in [v.upper() for v in VALID_VALUES['pf_cap']]:
                    pf_cap = pf_cap.upper()
                else:
                    pf_cap = DEFAULTS['pf_cap']

                esi_rule = ImportExportService._safe_str(row_data.get('esi_rule', ''))
                if esi_rule and esi_rule.upper() in [v.upper() for v in VALID_VALUES['esi_rule']]:
                    esi_rule = esi_rule.upper()
                else:
                    esi_rule = DEFAULTS['esi_rule']

                eps_applicable = ImportExportService._safe_str(row_data.get('eps_applicable', ''))
                if eps_applicable:
                    eps_applicable = eps_applicable.upper() in ['YES', 'TRUE', '1', 'Y']
                else:
                    eps_applicable = True

                pt_exempt = ImportExportService._safe_str(row_data.get('professional_tax_exempt', ''))
                if pt_exempt:
                    pt_exempt = pt_exempt.upper() in ['YES', 'TRUE', '1', 'Y']
                else:
                    pt_exempt = False

                status = ImportExportService._safe_str(row_data.get('status', ''))
                if status and status.upper() in [v.upper() for v in VALID_VALUES['status']]:
                    status = status.upper()
                else:
                    status = DEFAULTS['status']

                # ============================================
                # PARSE NUMBERS
                # ============================================
                monthly_basic = ImportExportService._parse_decimal(row_data.get('monthly_basic'))
                per_day_rate = ImportExportService._parse_decimal(row_data.get('per_day_rate'))
                rate_per_unit = ImportExportService._parse_decimal(row_data.get('rate_per_unit'))
                eps_employer_rate = ImportExportService._parse_decimal(
                    row_data.get('eps_employer_rate', DEFAULTS['eps_employer_rate']))
                eps_limit = ImportExportService._parse_decimal(row_data.get('eps_limit', DEFAULTS['eps_limit']))
                other_deductions = ImportExportService._parse_decimal(row_data.get('other_deductions', 0))

                # ============================================
                # COLLECT PAYHEAD VALUES
                # ============================================
                earnings = {}
                deductions = {}

                for payhead_name, col_idx in payhead_columns.items():
                    value = row_data.get(payhead_name)
                    if value is not None and str(value).strip():
                        try:
                            amount = float(str(value).strip())
                            if amount > 0:
                                clean_name = payhead_name.title().replace('_', ' ')
                                payhead_type = ImportExportService.get_payhead_type(payhead_name)

                                if payhead_type in ['EARNING', 'BONUS', 'ALLOWANCE', 'REIMBURSEMENT']:
                                    earnings[clean_name] = amount
                                else:
                                    deductions[clean_name] = amount
                        except ValueError:
                            pass

                salary_heads = {
                    'earnings': earnings,
                    'deductions': deductions
                }

                # ============================================
                # ✅ CREATE/UPDATE ASSIGNMENT ONLY
                # ============================================
                assignment = EmployeeAssignment.objects.filter(
                    employee=employee,
                    client=client,
                    location=location,
                    start_date=start_date
                ).first()

                if assignment:
                    assignment.end_date = end_date if end_date else assignment.end_date
                    assignment.effective_date = effective_date if effective_date else assignment.effective_date
                    assignment.salary_method = salary_method
                    assignment.monthly_basic = monthly_basic if monthly_basic is not None else assignment.monthly_basic
                    assignment.per_day_rate = per_day_rate if per_day_rate is not None else assignment.per_day_rate
                    assignment.rate_per_unit = rate_per_unit if rate_per_unit is not None else assignment.rate_per_unit
                    assignment.pf_cap = pf_cap
                    assignment.esi_rule = esi_rule
                    assignment.eps_applicable = eps_applicable
                    assignment.eps_employer_rate = eps_employer_rate if eps_employer_rate is not None else assignment.eps_employer_rate
                    assignment.eps_limit = eps_limit if eps_limit is not None else assignment.eps_limit
                    assignment.professional_tax_exempt = pt_exempt
                    assignment.other_deductions = other_deductions if other_deductions is not None else assignment.other_deductions
                    assignment.status = status
                    assignment.salary_heads = salary_heads
                    assignment.save()
                    update_count += 1
                else:
                    assignment = EmployeeAssignment.objects.create(
                        employee=employee,
                        client=client,
                        location=location,
                        start_date=start_date,
                        end_date=end_date,
                        effective_date=effective_date,
                        salary_method=salary_method,
                        monthly_basic=monthly_basic,
                        per_day_rate=per_day_rate,
                        rate_per_unit=rate_per_unit,
                        pf_cap=pf_cap,
                        esi_rule=esi_rule,
                        eps_applicable=eps_applicable,
                        eps_employer_rate=eps_employer_rate,
                        eps_limit=eps_limit,
                        professional_tax_exempt=pt_exempt,
                        other_deductions=other_deductions or Decimal('0'),
                        status=status,
                        salary_heads=salary_heads,
                    )
                    create_count += 1

                success_count += 1

            except Exception as e:
                errors.append({
                    'row': row_idx,
                    'column': 'General',
                    'error': str(e),
                    'value': ''
                })
                error_count += 1

        return {
            'success': True,
            'create_count': create_count,
            'update_count': update_count,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors,
            'created_payheads': created_payheads,
            'updated_payheads': updated_payheads,
        }

    # ============================================
    # EXPORT FUNCTIONS (unchanged)
    # ============================================

    @staticmethod
    def export_employees():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        # ... existing export code ...

    @staticmethod
    def get_employee_sample_template():
        """Generate sample Excel template for employees"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees Profile"

        # ... your existing code ...

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 40

        # ✅ ADD THIS - RETURN THE WORKBOOK
        return wb

    @staticmethod
    def import_clients(excel_file, user):
        # ... existing client import ...
        pass

    @staticmethod
    def export_clients():
        # ... existing client export ...
        pass

    @staticmethod
    def get_client_sample_template():
        # ... existing client template ...
        pass

    # apps/core/import_export.py - Add this method if missing

    @staticmethod
    def _export_errors_to_excel(errors):
        """Export errors to Excel file"""
        if not errors:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Errors"

        # Headers
        headers = ['Row Number', 'Column Name', 'Error Description', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Data
        for idx, error in enumerate(errors, 2):
            ws.cell(row=idx, column=1, value=error.get('row', ''))
            ws.cell(row=idx, column=2, value=error.get('column', ''))
            ws.cell(row=idx, column=3, value=error.get('error', ''))
            ws.cell(row=idx, column=4, value=error.get('value', ''))

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 30

        return wb

    # apps/core/import_export.py - Complete get_assignment_sample_template

    @staticmethod
    def get_assignment_sample_template():
        """Generate sample Excel template for Assignment Import with all Headings"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assignment Import"

        # Define colors
        REQUIRED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        EARNING_FILL = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        DEDUCTION_FILL = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        STATUTORY_FILL = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
        LOAN_FILL = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
        BONUS_FILL = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        ALLOWANCE_FILL = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")

        # Define border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ============================================
        # COMPLETE HEADERS - ALL COLUMNS (A to AE)
        # ============================================
        headers = [
            # Basic Fields (A-S)
            {'name': 'Employee Code', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Master Code', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Employee Name', 'required': False, 'color': HEADER_FILL, 'width': 25},
            {'name': 'Client Code', 'required': True, 'color': REQUIRED_FILL, 'width': 18},
            {'name': 'Location Code', 'required': True, 'color': REQUIRED_FILL, 'width': 18},
            {'name': 'Start Date', 'required': True, 'color': REQUIRED_FILL, 'width': 18},
            {'name': 'End Date', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Effective Date', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Salary Method', 'required': False, 'color': HEADER_FILL, 'width': 22},
            {'name': 'Monthly Basic', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Per Day Rate', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Rate Per Unit', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'PF Cap', 'required': False, 'color': HEADER_FILL, 'width': 20},
            {'name': 'ESI Rule', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'EPS Applicable', 'required': False, 'color': HEADER_FILL, 'width': 20},
            {'name': 'EPS Employer Rate', 'required': False, 'color': HEADER_FILL, 'width': 20},
            {'name': 'EPS Limit', 'required': False, 'color': HEADER_FILL, 'width': 18},
            {'name': 'Professional Tax Exempt', 'required': False, 'color': HEADER_FILL, 'width': 25},
            {'name': 'Other Deductions', 'required': False, 'color': HEADER_FILL, 'width': 20},
            {'name': 'Status', 'required': False, 'color': HEADER_FILL, 'width': 18},

            # Earnings (T-Z)
            {'name': 'Bonus', 'required': False, 'color': BONUS_FILL, 'width': 18},
            {'name': 'DA', 'required': False, 'color': EARNING_FILL, 'width': 18},
            {'name': 'HRA', 'required': False, 'color': ALLOWANCE_FILL, 'width': 18},
            {'name': 'Conveyance', 'required': False, 'color': ALLOWANCE_FILL, 'width': 20},
            {'name': 'Incentive', 'required': False, 'color': EARNING_FILL, 'width': 18},
            {'name': 'Special Allowance', 'required': False, 'color': ALLOWANCE_FILL, 'width': 22},
            {'name': 'Overtime', 'required': False, 'color': EARNING_FILL, 'width': 18},

            # Deductions (AA-AE)
            {'name': 'Insurance', 'required': False, 'color': DEDUCTION_FILL, 'width': 18},
            {'name': 'Advance', 'required': False, 'color': DEDUCTION_FILL, 'width': 18},
            {'name': 'Professional Tax', 'required': False, 'color': STATUTORY_FILL, 'width': 22},
            {'name': 'Loan Recovery', 'required': False, 'color': LOAN_FILL, 'width': 20},
        ]

        # ============================================
        # APPLY HEADERS
        # ============================================
        for col, header_info in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header_info['name'])
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_info['color']
            cell.border = thin_border

            # Set column width
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = header_info['width']

        # ============================================
        # ROW 2: HELP TEXT
        # ============================================
        help_texts = [
            'Employee Code (use with Master Code)',
            'Master Code (use with Employee Code)',
            'Employee Name (Display only)',
            'Client Code (must exist)',
            'Location Code (must exist under client)',
            'DD-MMM-YYYY, DD/MM/YYYY, YYYY-MM-DD',
            'DD-MMM-YYYY, DD/MM/YYYY, YYYY-MM-DD',
            'Defaults to Start Date',
            'CALENDAR_MONTH / 26_DAYS_MONTH / PER_DAY / PRODUCTION',
            'For Calendar Month method',
            'For Per Day method',
            'For Production method',
            'CAPPED_15000 / CAPPED_18000 / FULL / NOT_APPLICABLE',
            'AUTO / FORCE / EXEMPT',
            'YES/NO (Default: YES)',
            'Default: 8.33',
            'Default: 15000',
            'YES/NO (Default: NO)',
            'Fixed deduction amount',
            'ACTIVE / COMPLETED / TERMINATED / TRANSFERRED',
            'Earning - Bonus Amount',
            'Earning - Dearness Allowance',
            'Earning - House Rent Allowance',
            'Earning - Conveyance Allowance',
            'Earning - Incentive Amount',
            'Earning - Special Allowance',
            'Earning - Overtime Amount',
            'Deduction - Insurance Amount',
            'Deduction - Advance Amount',
            'Deduction - Professional Tax',
            'Deduction - Loan Recovery',
        ]

        for col, help_text in enumerate(help_texts, 1):
            cell = ws.cell(row=2, column=col, value=f"({help_text})")
            cell.font = Font(italic=True, size=8, color="666666")
            cell.alignment = Alignment(wrap_text=True)

        # ============================================
        # ROW 3: DEFAULTS & OPTIONS
        # ============================================
        default_texts = {
            'Salary Method': 'Options: CALENDAR_MONTH, 26_DAYS_MONTH, PER_DAY, PRODUCTION',
            'PF Cap': 'Options: CAPPED_15000, CAPPED_18000, FULL, NOT_APPLICABLE',
            'ESI Rule': 'Options: AUTO, FORCE, EXEMPT',
            'Status': 'Options: ACTIVE, COMPLETED, TERMINATED, TRANSFERRED',
            'EPS Applicable': 'Default: YES',
            'Professional Tax Exempt': 'Default: NO',
            'EPS Employer Rate': 'Default: 8.33',
            'EPS Limit': 'Default: 15000',
        }

        for col, header_info in enumerate(headers, 1):
            if header_info['name'] in default_texts:
                cell = ws.cell(row=3, column=col, value=default_texts[header_info['name']])
                cell.font = Font(italic=True, size=8, color="ED7D31")
                cell.alignment = Alignment(wrap_text=True)

        # ============================================
        # ROW 4: PAYHEAD TYPE INDICATORS
        # ============================================
        payhead_types = {
            'Bonus': '💰 BONUS',
            'DA': '💰 EARNING',
            'HRA': '💰 ALLOWANCE',
            'Conveyance': '💰 ALLOWANCE',
            'Incentive': '💰 EARNING',
            'Special Allowance': '💰 ALLOWANCE',
            'Overtime': '💰 EARNING',
            'Insurance': '💸 DEDUCTION',
            'Advance': '💸 DEDUCTION',
            'Professional Tax': '⚖️ STATUTORY',
            'Loan Recovery': '💳 LOAN',
        }

        for col, header_info in enumerate(headers, 1):
            if header_info['name'] in payhead_types:
                cell = ws.cell(row=4, column=col, value=payhead_types[header_info['name']])
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal="center")
                if 'BONUS' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="FFC107")
                elif 'EARNING' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="28A745")
                elif 'ALLOWANCE' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="17A2B8")
                elif 'DEDUCTION' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="DC3545")
                elif 'STATUTORY' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="FD7E14")
                elif 'LOAN' in payhead_types[header_info['name']]:
                    cell.font = Font(bold=True, size=8, color="6F42C1")

        # ============================================
        # SAMPLE DATA (Row 5 onwards)
        # ============================================
        sample_data = [
            # Employee Code, Master Code, Employee Name, Client Code, Location Code, Start Date, End Date, Effective Date,
            # Salary Method, Monthly Basic, Per Day Rate, Rate Per Unit,
            # PF Cap, ESI Rule, EPS Applicable, EPS Employer Rate, EPS Limit,
            # Professional Tax Exempt, Other Deductions, Status,
            # EARNINGS: Bonus, DA, HRA, Conveyance, Incentive, Special Allowance, Overtime,
            # DEDUCTIONS: Insurance, Advance, Professional Tax, Loan Recovery
            [
                'EMP0001', '', 'Rajesh Kumar', 'ABC', 'ABC-MUM', '05-Apr-2025', '31-Dec-2025', '',
                'CALENDAR_MONTH', '25000', '', '', 'CAPPED_15000', 'AUTO', 'YES', '8.33', '15000',
                'NO', '0', 'ACTIVE',
                '5000', '2000', '10000', '2000', '1000', '3000', '1500',
                '500', '2000', '200', '1000'
            ],
            [
                'EMP0002', '', 'Priya Sharma', 'ABC', 'ABC-MUM', '01-May-2025', '', '',
                'CALENDAR_MONTH', '30000', '', '', 'CAPPED_15000', 'AUTO', 'YES', '8.33', '15000',
                'NO', '0', 'ACTIVE',
                '', '', '12000', '', '', '', '',
                '', '', '', ''
            ],
            [
                '', 'EMP0003', 'Amit Singh', 'XYZ', 'XYZ-DEL', '15-May-2025', '', '',
                'PER_DAY', '', '2000', '', 'CAPPED_15000', 'AUTO', 'YES', '8.33', '15000',
                'NO', '0', 'ACTIVE',
                '3000', '', '', '1500', '', '', '',
                '1000', '', '', ''
            ],
            [
                'EMP0004', '', 'Sneha Patel', 'XYZ', 'XYZ-DEL', '01-Jun-2025', '', '',
                '26_DAYS_MONTH', '28000', '', '', 'FULL', 'FORCE', 'NO', '', '',
                'YES', '500', 'ACTIVE',
                '2000', '1500', '8000', '', '500', '', '',
                '', '1000', '300', ''
            ],
        ]

        for row_num, data in enumerate(sample_data, 5):
            for col_num, value in enumerate(data, 1):
                if col_num <= len(headers):
                    ws.cell(row=row_num, column=col_num, value=value)

        # ============================================
        # LEGEND / FOOTER
        # ============================================
        footer_row = len(sample_data) + 7

        # Legend
        ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
        legend_cell = ws.cell(row=footer_row, column=1,
                              value="📌 LEGEND: 🟥 Required Fields | 🟩 Earnings | 🟥 Deductions | 🟧 Statutory | 🟪 Loan | 🟨 Bonus | 🩵 Allowances")
        legend_cell.font = Font(bold=True, size=10)
        legend_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[footer_row].height = 25

        # Instructions
        instruction_row = footer_row + 1
        ws.merge_cells(f'A{instruction_row}:{get_column_letter(len(headers))}{instruction_row}')
        instruction_cell = ws.cell(row=instruction_row, column=1,
                                   value="📝 Employee Master Data (Bank, PAN, Aadhaar, Phone, ESI, UAN) is NOT modified during assignment import.")
        instruction_cell.font = Font(bold=True, size=10, color="0000FF")
        instruction_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[instruction_row].height = 30

        # Note
        note_row = instruction_row + 1
        ws.merge_cells(f'A{note_row}:{get_column_letter(len(headers))}{note_row}')
        note_cell = ws.cell(row=note_row, column=1,
                            value="💡 Employee Name is for display/reference only. Employee Code or Master Code is used to find employee.")
        note_cell.font = Font(italic=True, size=9, color="666666")
        note_cell.alignment = Alignment(horizontal="center")

        # ============================================
        # FREEZE PANES
        # ============================================
        ws.freeze_panes = 'A6'

        # ============================================
        # SET ROW HEIGHTS
        # ============================================
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 20

        # ✅ IMPORTANT: Return the workbook
        return wb